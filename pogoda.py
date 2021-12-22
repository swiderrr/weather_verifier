from selenium import webdriver
from openpyxl import Workbook, load_workbook
from datetime import date
from selenium.webdriver.common.by import By
import ctypes

start_date = date(2021, 12, 20)
today = date.today()

driver = webdriver.Firefox()
driver.get("https://www.meteoprog.pl/pl/weather/Warszawa/")

wb = Workbook()
path = 'pogoda.xlsx'
wb = load_workbook(path)
ws = wb.active

temperature_tommorow = driver.find_element(By.CLASS_NAME, 'weather-forecast__temperature')
temperature_today = driver.find_element(By.CLASS_NAME, 'today-temperature')

ws.set_column(1, 1, 150)
ws.set_column(2, 3, 100)

ws['A1'] = 'Data'
ws['B1'] = 'Realna temperatura'
ws['C1'] = 'Przewidywana temperatura'

ws['A' + str((today - start_date).days + 1)] = today
ws['B' + str((today - start_date).days + 1)] = temperature_today.text
ws['C' + str((today - start_date).days + 2)] = temperature_tommorow.text

ctypes.windll.user32.MessageBoxW(0, "Temperatura według wczorajszej prognozy to: {} \n Aktualna temperatura to: {}".format(ws['C' + str((today - start_date).days)].value, temperature_today.text), "Weather verificator", 0)

wb.save(path)
driver.close()
driver.quit()
